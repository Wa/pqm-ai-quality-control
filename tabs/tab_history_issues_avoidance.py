from __future__ import annotations

import hashlib
import io
import json
import os
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime
from typing import Callable, Iterable, Optional, Sequence

import pandas as pd
import requests
import streamlit as st
from ollama import Client as OllamaClient
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from config import CONFIG
from util import ensure_session_dirs, handle_file_upload, resolve_ollama_host
from bisheng_client import (
    call_flow_process,
    create_knowledge,
    find_knowledge_id_by_name,
    kb_sync_folder,
    parse_flow_answer,
)


PDF_EXTENSIONS = {".pdf"}
WORD_PPT_EXTENSIONS = {".doc", ".docx", ".ppt", ".pptx"}
SPREADSHEET_EXTENSIONS = {".xls", ".xlsx", ".xlsm", ".csv"}
HEADER_SCAN_LIMIT = 10
HEADER_SCORE_THRESHOLD = 4
LLM_CHUNK_LIMIT = 9500
HISTORY_KB_MODEL_ID = 7
HISTORY_FLOW_ID = "191af6f3565e415ca9670f1bc2b9117e"
HISTORY_FLOW_TWEAKS = {
    "MixEsVectorRetriever-J35CZ": {},
    "Milvus-cyR5W": {},
    "PromptTemplate-bs0vj": {},
    "BishengLLM-768ac": {},
    "ElasticKeywordsSearch-1c80e": {},
    "RetrievalQA-f0f31": {},
    "CombineDocsChain-2f68e": {},
}
HISTORY_FLOW_INPUT_NODE_ID = (
    os.getenv("HISTORY_FLOW_INPUT_NODE_ID")
    or os.getenv("FLOW_INPUT_NODE_ID")
    or "RetrievalQA-f0f31"
)
HISTORY_FLOW_MILVUS_NODE_ID = (
    os.getenv("HISTORY_FLOW_MILVUS_NODE_ID")
    or os.getenv("FLOW_MILVUS_NODE_ID")
    or "Milvus-cyR5W"
)
HISTORY_FLOW_ES_NODE_ID = (
    os.getenv("HISTORY_FLOW_ES_NODE_ID")
    or os.getenv("FLOW_ES_NODE_ID")
    or "ElasticKeywordsSearch-1c80e"
)


def _safe_int(value: str | None, fallback: int) -> int:
    try:
        return int(value) if value is not None else fallback
    except (TypeError, ValueError):
        return fallback


HISTORY_FLOW_TIMEOUT = _safe_int(
    os.getenv("HISTORY_FLOW_TIMEOUT")
    or os.getenv("BISHENG_TIMEOUT_S")
    or str(CONFIG.get("bisheng", {}).get("timeout_s", 120)),
    120,
)

HISTORY_FLOW_WARMUP_PROMPT = (
    os.getenv("HISTORY_FLOW_WARMUP_PROMPT") or "预热：请回复'历史问题比对准备就绪'。"
)

HEADER_ALIASES = {
    "failure_mode": [
        "失效模式",
        "问题",
        "问题描述",
        "历史问题",
        "不良现象",
        "故障模式",
        "Failure Mode",
    ],
    "root_cause": [
        "根因",
        "原因",
        "原因分析",
        "问题原因",
        "Root Cause",
        "发生原因",
    ],
    "prevention_action": [
        "预防措施",
        "预防行动",
        "改进措施",
        "控制措施",
        "永久措施",
        "Permanent Action",
    ],
    "detection_action": [
        "检测措施",
        "检测计划",
        "检验措施",
        "验证措施",
        "Detection",
    ],
    "severity": ["严重度", "严重性", "S"],
    "occurrence": ["发生度", "发生频度", "O"],
    "detection": ["探测度", "检测度", "D"],
    "risk_priority": ["RPN", "风险优先数", "风险评估"],
    "responsible": ["责任人", "负责人", "责任部门", "Owner"],
    "due_date": ["完成时间", "计划完成时间", "截止日期", "完成日期", "Due Date"],
    "status": ["状态", "进度", "落实情况"],
    "remarks": ["备注", "说明", "备注信息", "Remarks"],
}

REQUIRED_FIELDS = ["failure_mode", "root_cause", "prevention_action", "detection_action"]
OPTIONAL_FIELDS = [
    "severity",
    "occurrence",
    "detection",
    "risk_priority",
    "responsible",
    "due_date",
    "status",
    "remarks",
]

FULLWIDTH_TRANSLATION = str.maketrans({
    "，": ",",
    "。": ".",
    "；": ";",
    "：": ":",
    "？": "?",
    "！": "!",
    "（": "(",
    "）": ")",
    "【": "[",
    "】": "]",
    "“": '"',
    "”": '"',
    "‘": "'",
    "’": "'",
})

SYSTEM_PROMPT = (
    "你是一名资深质量工程师，需要从文本中提取历史问题记录。"
    "请把每段内容转换成JSON数组，每个元素至少包含"
    "failure_mode、root_cause、prevention_action、detection_action字段。"
    "如果无法找到某个字段，请使用空字符串。仅返回JSON，勿添加额外说明。"
)


class IssueRecord(BaseModel):
    failure_mode: str = Field(..., description="历史问题或失效模式")
    root_cause: str = ""
    prevention_action: str = ""
    detection_action: str = ""
    severity: str | None = None
    occurrence: str | None = None
    detection: str | None = None
    risk_priority: str | None = None
    responsible: str | None = None
    due_date: str | None = None
    status: str | None = None
    remarks: str | None = None
    hash_key: str = Field(..., description="同文件内用于去重的哈希")


@dataclass
class IssueSummary:
    total_rows: int = 0
    parsed_rows: int = 0
    skipped_rows: int = 0
    duplicates: int = 0

    def to_dict(self) -> dict[str, int]:
        return {
            "total_rows": self.total_rows,
            "parsed_rows": self.parsed_rows,
            "skipped_rows": self.skipped_rows,
            "duplicates": self.duplicates,
        }


def _list_files(folder: str, extensions: Iterable[str] | None = None) -> list[str]:
    if not folder or not os.path.isdir(folder):
        return []
    results: list[str] = []
    for name in os.listdir(folder):
        path = os.path.join(folder, name)
        if not os.path.isfile(path):
            continue
        if extensions:
            ext = os.path.splitext(name)[1].lower()
            if ext not in extensions:
                continue
        results.append(path)
    return results


def _list_pdfs(folder: str) -> list[str]:
    return _list_files(folder, PDF_EXTENSIONS)


def _mineru_parse_pdf(pdf_path: str) -> bytes:
    api_url = "http://10.31.60.127:8000/file_parse"
    data = {
        "backend": "vlm-sglang-engine",
        "response_format_zip": "true",
        "formula_enable": "true",
        "table_enable": "true",
        "return_images": "false",
        "return_middle_json": "true",
        "return_model_output": "false",
        "return_content_list": "true",
    }
    with open(pdf_path, "rb") as f:
        files = {"files": (os.path.basename(pdf_path), f, "application/pdf")}
        resp = requests.post(api_url, data=data, files=files, timeout=300)
        if resp.status_code != 200:
            raise RuntimeError(f"MinerU API error {resp.status_code}: {resp.text[:200]}")
        return resp.content


def _zip_to_txts(zip_bytes: bytes, target_txt_path: str) -> bool:
    bio = io.BytesIO(zip_bytes)
    try:
        with zipfile.ZipFile(bio) as zf:
            md_members = [n for n in zf.namelist() if n.lower().endswith(".md")]
            if not md_members:
                return False
            name = md_members[0]
            content = zf.read(name)
            os.makedirs(os.path.dirname(target_txt_path), exist_ok=True)
            with open(target_txt_path, "wb") as out_f:
                out_f.write(content)
            return True
    except zipfile.BadZipFile:
        return False


def _list_word_ppt(folder: str) -> list[str]:
    return _list_files(folder, WORD_PPT_EXTENSIONS)


def _unstructured_partition_to_txt(file_path: str, target_txt_path: str) -> bool:
    api_url = (
        os.getenv("UNSTRUCTURED_API_URL")
        or CONFIG.get("services", {}).get("unstructured_api_url")
        or "http://10.31.60.11:8000/general/v0/general"
    )
    try:
        with open(file_path, "rb") as f:
            files = {"files": (os.path.basename(file_path), f)}
            form = {
                "strategy": "auto",
                "ocr_languages": "chi_sim,eng",
                "infer_table_structure": "true",
            }
            resp = requests.post(api_url, files=files, data=form, timeout=300)
            if resp.status_code != 200:
                raise RuntimeError(f"Unstructured API {resp.status_code}: {resp.text[:200]}")
            data = resp.json()
            lines: list[str] = []
            if isinstance(data, list):
                for el in data:
                    text = None
                    if isinstance(el, dict):
                        text = el.get("text")
                        if not text and isinstance(el.get("data"), list):
                            for row in el["data"]:
                                if isinstance(row, list):
                                    lines.append("\t".join(str(c) for c in row))
                            continue
                    if isinstance(text, str) and text.strip():
                        lines.append(text.strip())
            os.makedirs(os.path.dirname(target_txt_path), exist_ok=True)
            with open(target_txt_path, "w", encoding="utf-8") as out_f:
                out_f.write("\n".join(lines))
            return True
    except Exception:
        return False
    return False


def _list_excels(folder: str) -> list[str]:
    return _list_files(folder, SPREADSHEET_EXTENSIONS - {".csv"})


def _collect_files(folder: str) -> list[dict[str, object]]:
    if not folder or not os.path.isdir(folder):
        return []
    items = []
    for name in os.listdir(folder):
        path = os.path.join(folder, name)
        if not os.path.isfile(path):
            continue
        stat = os.stat(path)
        items.append(
            {
                "name": name,
                "path": path,
                "size": stat.st_size,
                "modified": stat.st_mtime,
            }
        )
    items.sort(key=lambda info: info["modified"], reverse=True)
    return items


def _format_file_size(size_bytes: int) -> str:
    if size_bytes == 0:
        return "0 B"
    units = ["B", "KB", "MB", "GB"]
    value = float(size_bytes)
    index = 0
    while value >= 1024 and index < len(units) - 1:
        value /= 1024.0
        index += 1
    return f"{value:.1f} {units[index]}"


def _format_timestamp(timestamp: float) -> str:
    return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M")


def _truncate_filename(filename: str, max_length: int = 40) -> str:
    if len(filename) <= max_length:
        return filename
    name, ext = os.path.splitext(filename)
    available = max_length - len(ext) - 3
    if available <= 0:
        return filename[: max_length - 3] + "..."
    return name[:available] + "..." + ext


def _ensure_generated_dirs(root: str) -> dict[str, str]:
    mapping: dict[str, str] = {}
    os.makedirs(root, exist_ok=True)
    for key in ("issue_lists", "dfmea", "pfmea", "cp"):
        path = os.path.join(root, f"{key}_txt")
        os.makedirs(path, exist_ok=True)
        mapping[key] = path
    initial_dir = os.path.join(root, "initial_results")
    final_dir = os.path.join(root, "final_results")
    os.makedirs(initial_dir, exist_ok=True)
    os.makedirs(final_dir, exist_ok=True)
    mapping["initial_results"] = initial_dir
    mapping["final_results"] = final_dir
    return mapping


def _sanitize_sheet_name(name: str) -> str:
    bad = ["\\", "/", ":", "*", "?", '"', "<", ">", "|"]
    for ch in bad:
        name = name.replace(ch, "_")
    return "_".join(name.strip().split())[:80] or "Sheet"


def _process_pdf_folder(input_dir: str, output_dir: str, progress_area) -> list[str]:
    pdf_paths = _list_pdfs(input_dir)
    created: list[str] = []
    if not pdf_paths:
        progress_area.info("（无PDF文件可处理）")
        return created
    for pdf_path in pdf_paths:
        orig_name = os.path.basename(pdf_path)
        out_txt = os.path.join(output_dir, f"{orig_name}.txt")
        try:
            if os.path.exists(out_txt) and os.path.getsize(out_txt) > 0:
                progress_area.info(f"已存在（跳过）: {os.path.basename(out_txt)}")
                continue
            progress_area.write(f"解析PDF: {orig_name} …")
            zip_bytes = _mineru_parse_pdf(pdf_path)
            ok = _zip_to_txts(zip_bytes, out_txt)
            if ok:
                created.append(out_txt)
                progress_area.success(f"已生成: {os.path.basename(out_txt)}")
            else:
                progress_area.warning(f"未生成文本，跳过: {orig_name}")
        except Exception as error:
            progress_area.error(f"失败: {orig_name} → {error}")
    return created


def _process_word_ppt_folder(input_dir: str, output_dir: str, progress_area) -> list[str]:
    doc_paths = _list_word_ppt(input_dir)
    created: list[str] = []
    if not doc_paths:
        progress_area.info("（无Word/PPT文件可处理）")
        return created
    for file_path in doc_paths:
        orig_name = os.path.basename(file_path)
        out_txt = os.path.join(output_dir, f"{orig_name}.txt")
        try:
            if os.path.exists(out_txt) and os.path.getsize(out_txt) > 0:
                progress_area.info(f"已存在（跳过）: {os.path.basename(out_txt)}")
                continue
            progress_area.write(f"解析文档: {orig_name} …")
            ok = _unstructured_partition_to_txt(file_path, out_txt)
            if ok:
                created.append(out_txt)
                progress_area.success(f"已生成: {os.path.basename(out_txt)}")
            else:
                progress_area.warning(f"未生成文本，跳过: {orig_name}")
        except Exception as error:
            progress_area.error(f"失败: {orig_name} → {error}")
    return created


def _process_excel_folder(input_dir: str, output_dir: str, progress_area) -> list[str]:
    excel_paths = _list_excels(input_dir)
    created: list[str] = []
    if not excel_paths:
        progress_area.info("（无Excel文件可处理）")
        return created
    for excel_path in excel_paths:
        orig_name = os.path.basename(excel_path)
        try:
            xls = pd.ExcelFile(excel_path)
            for sheet in xls.sheet_names:
                safe_sheet = _sanitize_sheet_name(sheet)
                out_txt = os.path.join(output_dir, f"{orig_name}_SHEET_{safe_sheet}.txt")
                if os.path.exists(out_txt) and os.path.getsize(out_txt) > 0:
                    progress_area.info(f"已存在（跳过）: {os.path.basename(out_txt)}")
                    continue
                progress_area.write(f"转换Excel: {orig_name} / {sheet} …")
                df = xls.parse(sheet)
                df.to_csv(out_txt, index=False, encoding="utf-8")
                created.append(out_txt)
                progress_area.success(f"已生成: {os.path.basename(out_txt)}")
        except Exception as error:
            progress_area.error(f"失败: {orig_name} → {error}")
    return created


def _process_category(label: str, source_dir: str | None, output_dir: str, progress_area) -> list[str]:
    os.makedirs(output_dir, exist_ok=True)
    if not source_dir or not os.path.isdir(source_dir):
        progress_area.warning(f"未找到 {label} 上传目录，已跳过。")
        return []
    progress_area.markdown(f"**{label} → 文本转换**")
    created: list[str] = []
    created.extend(_process_pdf_folder(source_dir, output_dir, progress_area))
    created.extend(_process_word_ppt_folder(source_dir, output_dir, progress_area))
    created.extend(_process_excel_folder(source_dir, output_dir, progress_area))
    if not created:
        progress_area.info(f"{label} 未生成任何文本文件，请确认已上传 PDF/Word/Excel。")
    return created


def _safe_result_name(base: str) -> str:
    name = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff._-]", "_", base)
    name = name.strip("._")
    return name[:120] or "result"


def _sanitize_bisheng_base_url(raw: str | None) -> str:
    if not raw:
        return ""
    trimmed = raw.strip()
    if not trimmed:
        return ""
    trimmed = trimmed.rstrip("/")
    suffixes = [
        "/api/v1/process",
        "/api/v2/workflow/invoke",
        "/api/v2/workflow",
        "/api/v2",
    ]
    lowered = trimmed.lower()
    for suffix in suffixes:
        if lowered.endswith(suffix):
            trimmed = trimmed[: -len(suffix)].rstrip("/")
            lowered = trimmed.lower()
    return trimmed


def _resolve_bisheng_credentials() -> tuple[str, Optional[str]]:
    settings = CONFIG.get("bisheng", {})
    base_url = (
        os.getenv("HISTORY_BISHENG_BASE_URL")
        or os.getenv("HISTORY_FLOW_BASE_URL")
        or os.getenv("HISTORY_FLOW_BASE")
        or os.getenv("BISHENG_BASE_URL")
        or settings.get("base_url")
        or ""
    )
    api_key = (
        os.getenv("HISTORY_BISHENG_API_KEY")
        or os.getenv("BISHENG_API_KEY")
        or settings.get("api_key")
        or ""
    )
    return _sanitize_bisheng_base_url(base_url), (api_key or None)


def _build_history_kb_name(session_id: str) -> str:
    base = f"{session_id}_history"
    return _safe_result_name(base)


def _sync_history_kb(
    session_id: str,
    text_dirs: dict[str, str],
    progress_area,
) -> Optional[int]:
    base_url, api_key = _resolve_bisheng_credentials()
    if not base_url:
        progress_area.warning("未配置毕昇服务地址，跳过知识库同步。")
        return None
    kb_name = _build_history_kb_name(session_id)
    try:
        knowledge_id = find_knowledge_id_by_name(base_url, api_key, kb_name)
        if knowledge_id:
            progress_area.info(f"已找到知识库：{kb_name} (ID: {knowledge_id})")
        else:
            progress_area.write(f"正在创建知识库：{kb_name} …")
            knowledge_id = create_knowledge(
                base_url,
                api_key,
                kb_name,
                model=str(HISTORY_KB_MODEL_ID),
                description="历史问题规避-项目文档",
            )
            if knowledge_id:
                progress_area.success(f"已创建知识库：{kb_name} (ID: {knowledge_id})")
        if not knowledge_id:
            progress_area.warning(
                "无法创建或获取知识库，跳过同步。请检查毕昇服务配置与权限。"
            )
            return None
        total_uploaded = 0
        total_skipped = 0
        for label, folder in text_dirs.items():
            if not folder or not os.path.isdir(folder):
                continue
            progress_area.write(f"同步至知识库 ({label}) …")
            try:
                result = kb_sync_folder(
                    base_url=base_url,
                    api_key=api_key,
                    knowledge_id=int(knowledge_id),
                    folder_path=folder,
                    clear_first=False,
                    chunk_size=800,
                    chunk_overlap=80,
                    separators=["\n\n", "\n"],
                    separator_rule=["after", "after"],
                )
            except Exception as error:
                progress_area.error(f"同步 {label} 失败: {error}")
                continue
            uploaded = len(result.get("uploaded", [])) if isinstance(result, dict) else 0
            skipped = len(result.get("skipped", [])) if isinstance(result, dict) else 0
            total_uploaded += uploaded
            total_skipped += skipped
            progress_area.success(
                f"{label} 已同步：上传 {uploaded} 个，跳过 {skipped} 个。"
            )
        if total_uploaded or total_skipped:
            progress_area.info(
                f"知识库同步完成（上传 {total_uploaded}，跳过 {total_skipped}）。"
            )
        else:
            progress_area.info("知识库同步完成（无可同步文件）。")
        return int(knowledge_id)
    except Exception as error:
        progress_area.error(f"知识库同步失败: {error}")
        return None


def _apply_kb_to_tweaks(kb_id: Optional[int]) -> dict:
    tweaks = {key: value.copy() for key, value in HISTORY_FLOW_TWEAKS.items()}
    if kb_id is None:
        return tweaks
    for node_id in (HISTORY_FLOW_MILVUS_NODE_ID, HISTORY_FLOW_ES_NODE_ID):
        node_tw = tweaks.get(node_id, {}).copy()
        node_tw["collection_id"] = str(kb_id)
        tweaks[node_id] = node_tw
    return tweaks


def _warmup_history_flow(kb_id: Optional[int], progress_area) -> None:
    base_url, api_key = _resolve_bisheng_credentials()
    if not base_url:
        progress_area.warning("未配置毕昇服务地址，跳过流程预热。")
        return
    tweaks = _apply_kb_to_tweaks(kb_id)
    try:
        call_flow_process(
            base_url=base_url,
            flow_id=HISTORY_FLOW_ID,
            question=HISTORY_FLOW_WARMUP_PROMPT,
            kb_id=kb_id,
            input_node_id=HISTORY_FLOW_INPUT_NODE_ID,
            api_key=api_key,
            session_id=None,
            history_count=0,
            extra_tweaks=tweaks,
            milvus_node_id=HISTORY_FLOW_MILVUS_NODE_ID,
            es_node_id=HISTORY_FLOW_ES_NODE_ID,
            timeout_s=min(30, HISTORY_FLOW_TIMEOUT),
            max_retries=0,
        )
        progress_area.info("毕昇比对流程预热完成。")
    except Exception as error:
        progress_area.warning(f"毕昇比对流程预热失败：{error}")


def _invoke_history_flow(
    prompt: str,
    kb_id: Optional[int],
    session_token: Optional[str],
    progress_area,
) -> dict:
    base_url, api_key = _resolve_bisheng_credentials()
    if not base_url:
        progress_area.error("未配置毕昇服务地址，无法调用比对流程。")
        return {"error": "missing_bisheng_base_url"}
    tweaks = _apply_kb_to_tweaks(kb_id)
    try:
        response = call_flow_process(
            base_url=base_url,
            flow_id=HISTORY_FLOW_ID,
            question=prompt,
            kb_id=kb_id,
            input_node_id=HISTORY_FLOW_INPUT_NODE_ID,
            api_key=api_key,
            session_id=session_token,
            history_count=0,
            extra_tweaks=tweaks,
            milvus_node_id=HISTORY_FLOW_MILVUS_NODE_ID,
            es_node_id=HISTORY_FLOW_ES_NODE_ID,
            timeout_s=HISTORY_FLOW_TIMEOUT,
            max_retries=2,
        )
        return response
    except Exception as error:
        progress_area.error(f"调用毕昇比对失败: {error}")
        return {"error": str(error)}


def _build_record_prompt(record: IssueRecord) -> str:
    details = [
        f"失效模式: {record.failure_mode}",
        f"根因: {record.root_cause or '（未提供）'}",
        f"预防措施: {record.prevention_action or '（未提供）'}",
        f"检测措施: {record.detection_action or '（未提供）'}",
    ]
    for label, value in (
        ("严重度", record.severity),
        ("发生度", record.occurrence),
        ("检测度", record.detection),
        ("责任人", record.responsible),
        ("计划完成时间", record.due_date),
        ("状态", record.status),
        ("备注", record.remarks),
    ):
        if value:
            details.append(f"{label}: {value}")
    instructions = """
你是一名资深APQP质量工程师，负责检查历史问题是否已在当前项目的DFMEA、PFMEA、控制计划中得到预防。

你的任务：
1. 阅读“历史问题”的描述，理解其失效模式、根本原因、预防措施、检测措施。
2. 阅读提供的当前项目文档片段（DFMEA/PFMEA/控制计划）。
3. 判断该历史问题的预防与检测措施是否已经被覆盖。
4. 若未覆盖，请建议在哪个文档（DFMEA/PFMEA/控制计划）中增加何种控制。

请仅输出以下JSON格式：
{
  "status": "已覆盖 | 部分覆盖 | 未覆盖",
  "where_covered": [
    {"doc_type": "PFMEA", "row_ref": "PFMEA-R12", "说明": "已有UV固化时间5s及拉力测试控制"}
  ],
  "建议更新": [
    {"目标文件": "控制计划", "建议内容": "增加100%拉力测试≥5N", "理由": "对应历史问题NTC探头松脱"}
  ]
}
""".strip()
    return f"{instructions}\n\n历史问题详情：\n" + "\n".join(f"- {line}" for line in details if line)


def _load_issue_payloads(issue_dir: str) -> list[tuple[str, list[IssueRecord]]]:
    payloads: list[tuple[str, list[IssueRecord]]] = []
    if not issue_dir or not os.path.isdir(issue_dir):
        return payloads
    for name in sorted(os.listdir(issue_dir)):
        path = os.path.join(issue_dir, name)
        if not os.path.isfile(path) or not name.lower().endswith(".txt"):
            continue
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            continue
        records_raw = data.get("records") if isinstance(data, dict) else None
        if not isinstance(records_raw, list):
            continue
        records: list[IssueRecord] = []
        for item in records_raw:
            if not isinstance(item, dict):
                continue
            try:
                records.append(IssueRecord(**item))
            except Exception:
                continue
        if records:
            payloads.append((name, records))
    return payloads


def _evaluate_history_problems(
    issue_dir: str,
    initial_dir: str,
    final_dir: str,
    kb_id: Optional[int],
    session_id: str,
    progress_area,
    transcript_area=None,
    progress_callback: Optional[Callable[[float], None]] = None,
    progress_offset: float = 0.0,
    progress_span: float = 0.85,
) -> None:
    os.makedirs(initial_dir, exist_ok=True)
    os.makedirs(final_dir, exist_ok=True)
    payloads = _load_issue_payloads(issue_dir)
    if progress_callback:
        progress_callback(progress_offset)
    if not payloads:
        progress_area.info("未发现历史问题解析结果，跳过比对。")
        if progress_callback:
            progress_callback(min(progress_offset + progress_span, 1.0))
        return
    total_records = sum(len(records) for _, records in payloads)
    if total_records == 0:
        progress_area.info("历史问题记录为空。")
        if progress_callback:
            progress_callback(min(progress_offset + progress_span, 1.0))
        return
    progress_area.markdown("**开始历史问题覆盖性比对**")
    collected_rows: list[dict[str, object]] = []
    processed = 0
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    flow_session_key = f"history_bisheng_session_{session_id}"
    for file_name, records in payloads:
        safe_file = _safe_result_name(os.path.splitext(file_name)[0])
        for index, record in enumerate(records, start=1):
            prompt = _build_record_prompt(record)
            session_token = st.session_state.get(flow_session_key)
            response = _invoke_history_flow(prompt, kb_id, session_token, progress_area)
            answer_text, response_session = parse_flow_answer(response)
            if response_session:
                st.session_state[flow_session_key] = response_session
            parsed_answer: dict[str, object] | None = None
            if isinstance(answer_text, str) and answer_text.strip():
                try:
                    parsed_answer = json.loads(answer_text)
                except json.JSONDecodeError:
                    parsed_answer = None
            initial_payload = {
                "source_file": file_name,
                "record_index": index,
                "prompt": prompt,
                "record": record.model_dump(),
                "response": response,
                "answer": answer_text,
            }
            out_name = f"{safe_file}_record_{index:04d}.json"
            try:
                with open(os.path.join(initial_dir, out_name), "w", encoding="utf-8") as f:
                    json.dump(initial_payload, f, ensure_ascii=False, indent=2)
            except Exception as error:
                progress_area.warning(f"写入初步结果失败 ({out_name}): {error}")
            if transcript_area is not None:
                transcript_area.markdown(
                    f"**{file_name} · 第 {index} 条记录**"
                )
                transcript_area.markdown("**提示词**")
                transcript_area.code(prompt, language="markdown")
                transcript_area.markdown("**回答**")
                transcript_area.write(answer_text or "（无回答）")
                transcript_area.divider()
            row = {
                "source_file": file_name,
                "record_index": index,
                "session_id": session_id,
                "failure_mode": record.failure_mode,
                "root_cause": record.root_cause,
                "prevention_action": record.prevention_action,
                "detection_action": record.detection_action,
                "status": "",
                "where_covered": "",
                "建议更新": "",
                "原始回答": answer_text or "",
            }
            if parsed_answer:
                row["status"] = str(parsed_answer.get("status", ""))
                row["where_covered"] = json.dumps(
                    parsed_answer.get("where_covered", []), ensure_ascii=False
                )
                row["建议更新"] = json.dumps(
                    parsed_answer.get("建议更新", []), ensure_ascii=False
                )
            collected_rows.append(row)
            processed += 1
            if progress_callback and total_records:
                fraction = processed / total_records
                progress_callback(min(progress_offset + progress_span * fraction, 1.0))
    if not collected_rows:
        progress_area.info("未获取到有效的比对结果。")
        if progress_callback:
            progress_callback(min(progress_offset + progress_span, 1.0))
        return
    df = pd.DataFrame(collected_rows)
    csv_path = os.path.join(final_dir, f"history_issues_results_{timestamp}.csv")
    xlsx_path = os.path.join(final_dir, f"history_issues_results_{timestamp}.xlsx")
    try:
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        progress_area.success(f"已生成CSV结果：{os.path.basename(csv_path)}")
    except Exception as error:
        progress_area.error(f"导出CSV失败: {error}")
    try:
        df.to_excel(xlsx_path, index=False, engine="openpyxl")
        progress_area.success(f"已生成Excel结果：{os.path.basename(xlsx_path)}")
    except Exception as error:
        progress_area.error(f"导出Excel失败: {error}")
    if progress_callback:
        progress_callback(min(progress_offset + progress_span, 1.0))


def _strip_html(value: str) -> str:
    return re.sub(r"<[^>]+>", "", value)


def _normalize_text(value: str) -> str:
    text = value.replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.translate(FULLWIDTH_TRANSLATION).strip()


def _clean_cell_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return _normalize_text(_strip_html(value))
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime("%Y-%m-%d")
    return _normalize_text(str(value))


def _horizontal_fill(row: Sequence[str]) -> list[str]:
    filled: list[str] = []
    last = ""
    for cell in row:
        text = cell or ""
        if not text and last:
            filled.append(last)
        else:
            filled.append(text)
            if text:
                last = text
    return filled


def _build_header_candidate(rows: list[list[str]], indices: Sequence[int]) -> list[str]:
    selected = []
    for idx in indices:
        if idx < len(rows):
            selected.append(_horizontal_fill(rows[idx]))
    if not selected:
        return []
    width = max(len(row) for row in selected)
    headers: list[str] = []
    for col in range(width):
        parts = []
        for row in selected:
            value = row[col] if col < len(row) else ""
            if value:
                parts.append(value)
        headers.append("/".join(parts).strip("/"))
    return headers


def _match_header(header: str) -> tuple[str | None, int]:
    if not header:
        return None, 0
    normalized = header.lower()
    best_key = None
    best_score = 0
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_norm = alias.lower()
            if normalized == alias_norm:
                score = 100 + len(alias_norm)
            elif alias_norm in normalized or normalized in alias_norm:
                score = len(alias_norm)
            else:
                continue
            if score > best_score:
                best_score = score
                best_key = canonical
    return best_key, best_score


def map_headers(headers: Sequence[str]) -> tuple[dict[int, str], int]:
    mapping: dict[int, str] = {}
    total_score = 0
    used: set[str] = set()
    for index, header in enumerate(headers):
        canonical, score = _match_header(header)
        if canonical and (canonical not in used or score >= 100):
            mapping[index] = canonical
            total_score += score
            used.add(canonical)
    return mapping, total_score


def _detect_header(rows: list[list[str]]) -> tuple[list[int], dict[int, str]]:
    candidates: list[tuple[int, list[int], dict[int, str]]] = []
    scan_limit = min(HEADER_SCAN_LIMIT, len(rows))
    for row_index in range(scan_limit):
        headers = _build_header_candidate(rows, [row_index])
        mapping, score = map_headers(headers)
        if score >= HEADER_SCORE_THRESHOLD and "failure_mode" in mapping.values():
            candidates.append((score, [row_index], mapping))
    for row_index in range(scan_limit - 1):
        combo = [row_index, row_index + 1]
        headers = _build_header_candidate(rows, combo)
        mapping, score = map_headers(headers)
        if score >= HEADER_SCORE_THRESHOLD and "failure_mode" in mapping.values():
            candidates.append((score, combo, mapping))
    if not candidates:
        raise ValueError("未能识别表头")
    candidates.sort(key=lambda item: (item[0], -len(item[1])), reverse=True)
    _, indices, mapping = candidates[0]
    return indices, mapping


def _prepare_rows(raw_rows: list[list]) -> list[list[str]]:
    processed: list[list[str]] = []
    for raw_row in raw_rows:
        processed.append([_clean_cell_value(cell) for cell in raw_row])
    return processed


def _iter_excel_rows(workbook_path: str) -> Iterable[tuple[str, list[list]]]:
    wb = load_workbook(workbook_path, data_only=True, read_only=False)
    for sheet in wb.worksheets:
        if sheet.sheet_state != "visible":
            continue
        rows: list[list] = []
        for row in sheet.iter_rows(values_only=True):
            if row is None:
                continue
            rows.append(list(row))
        if all(not any(cell is not None and str(cell).strip() for cell in row) for row in rows):
            continue
        yield sheet.title, rows


def _load_csv_rows(path: str) -> list[list]:
    encodings = ["utf-8-sig", "utf-8", "gbk", "gb2312"]
    for encoding in encodings:
        try:
            df = pd.read_csv(
                path,
                header=None,
                dtype=str,
                keep_default_na=False,
                encoding=encoding,
            )
            return df.fillna("").values.tolist()
        except Exception:
            continue
    raise ValueError("无法解析CSV编码")


def _build_issue_record(row_map: dict[str, str]) -> IssueRecord | None:
    failure_mode = row_map.get("failure_mode", "").strip()
    if not failure_mode:
        return None
    hash_input = "|".join(
        row_map.get(field, "").strip().lower() for field in REQUIRED_FIELDS
    )
    hash_key = hashlib.sha1(hash_input.encode("utf-8")).hexdigest()
    record_data: dict[str, str | None] = {
        "failure_mode": failure_mode,
        "root_cause": row_map.get("root_cause", ""),
        "prevention_action": row_map.get("prevention_action", ""),
        "detection_action": row_map.get("detection_action", ""),
        "hash_key": hash_key,
    }
    for field in OPTIONAL_FIELDS:
        record_data[field] = row_map.get(field) or None
    return IssueRecord(**record_data)


def _parse_issue_sheet(rows: list[list[str]]) -> tuple[list[IssueRecord], IssueSummary]:
    rows_prepared = [_horizontal_fill(row) for row in rows]
    header_indices, mapping = _detect_header(rows_prepared)
    data_start = max(header_indices) + 1
    width = max(len(row) for row in rows_prepared) if rows_prepared else 0
    summary = IssueSummary()
    records: list[IssueRecord] = []
    seen: set[str] = set()
    prev_values = ["" for _ in range(width)]
    for row_idx in range(data_start, len(rows_prepared)):
        raw_row = rows_prepared[row_idx]
        values = [_clean_cell_value(raw_row[col]) if col < len(raw_row) else "" for col in range(width)]
        for col, value in enumerate(values):
            if not value and prev_values[col]:
                values[col] = prev_values[col]
            else:
                prev_values[col] = value
        summary.total_rows += 1
        row_map: dict[str, str] = {}
        for col_index, canonical in mapping.items():
            value = values[col_index] if col_index < len(values) else ""
            row_map[canonical] = value
        record = _build_issue_record(row_map)
        if record is None:
            summary.skipped_rows += 1
            continue
        if record.hash_key in seen:
            summary.duplicates += 1
            continue
        seen.add(record.hash_key)
        records.append(record)
        summary.parsed_rows += 1
    return records, summary


def _parse_issue_list_spreadsheet(file_path: str) -> tuple[list[IssueRecord], IssueSummary]:
    ext = os.path.splitext(file_path)[1].lower()
    all_records: list[IssueRecord] = []
    summary = IssueSummary()
    if ext == ".csv":
        rows = _prepare_rows(_load_csv_rows(file_path))
        records, part_summary = _parse_issue_sheet(rows)
        summary.total_rows += part_summary.total_rows
        summary.parsed_rows += part_summary.parsed_rows
        summary.skipped_rows += part_summary.skipped_rows
        summary.duplicates += part_summary.duplicates
        all_records.extend(records)
        return all_records, summary
    for sheet_name, raw_rows in _iter_excel_rows(file_path):
        rows = _prepare_rows(raw_rows)
        try:
            records, part_summary = _parse_issue_sheet(rows)
            summary.total_rows += part_summary.total_rows
            summary.parsed_rows += part_summary.parsed_rows
            summary.skipped_rows += part_summary.skipped_rows
            summary.duplicates += part_summary.duplicates
            if not records:
                continue
            all_records.extend(records)
        except ValueError as error:
            raise ValueError(f"{sheet_name}: {error}") from error
    return all_records, summary


def _chunk_text(text: str, limit: int = LLM_CHUNK_LIMIT) -> list[str]:
    if len(text) <= limit:
        return [text]
    segments: list[str] = []
    current: list[str] = []
    current_len = 0
    for line in text.splitlines(keepends=True):
        line_len = len(line)
        if current and current_len + line_len > limit:
            segments.append("".join(current))
            current = [line]
            current_len = line_len
        else:
            current.append(line)
            current_len += line_len
    if current:
        segments.append("".join(current))
    return segments


def _call_llm_for_issue_lists(text: str, file_label: str, progress_area) -> tuple[list[IssueRecord], IssueSummary]:
    host = resolve_ollama_host("ollama_9")
    try:
        client = OllamaClient(host=host)
    except Exception as error:
        progress_area.error(f"无法连接 gpt-oss: {error}")
        return [], IssueSummary()
    model = CONFIG["llm"].get("ollama_model", "gpt-oss:latest")
    segments = _chunk_text(text, LLM_CHUNK_LIMIT)
    aggregated: list[IssueRecord] = []
    summary = IssueSummary()
    seen: set[str] = set()
    for index, segment in enumerate(segments, start=1):
        prompt = (
            f"以下是历史问题清单文本（第{index}/{len(segments)}段，文件: {file_label}）。"
            "请提取问题条目并返回JSON数组。"
            "字段必须包含failure_mode、root_cause、prevention_action、detection_action，"
            "其余可根据内容补充，可选字段使用空字符串。"
            "禁止输出除JSON外的任何文本。\n\n" + segment
        )
        try:
            result = client.chat(
                model=model,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user", "content": prompt},
                ],
                options={"temperature": 0.1},
            )
        except Exception as error:
            progress_area.error(f"调用 gpt-oss 失败 ({file_label} 第{index}段): {error}")
            continue
        content = ""
        if isinstance(result, dict):
            content = result.get("message", {}).get("content") or result.get("response") or ""
        if not content:
            progress_area.warning(f"gpt-oss 未返回内容 ({file_label} 第{index}段)")
            continue
        try:
            parsed = json.loads(content)
        except json.JSONDecodeError as error:
            progress_area.error(f"gpt-oss 返回内容无法解析JSON ({file_label} 第{index}段): {error}")
            continue
        if isinstance(parsed, dict) and "records" in parsed:
            parsed = parsed.get("records")
        if not isinstance(parsed, list):
            progress_area.warning(f"gpt-oss 返回格式异常 ({file_label} 第{index}段)")
            continue
        summary.total_rows += len(parsed)
        for item in parsed:
            if not isinstance(item, dict):
                summary.skipped_rows += 1
                continue
            normalized = {
                "failure_mode": _clean_cell_value(item.get("failure_mode")),
                "root_cause": _clean_cell_value(item.get("root_cause")),
                "prevention_action": _clean_cell_value(item.get("prevention_action")),
                "detection_action": _clean_cell_value(item.get("detection_action")),
            }
            for field in OPTIONAL_FIELDS:
                normalized[field] = _clean_cell_value(item.get(field)) if field in item else None
            record = _build_issue_record(normalized)
            if record is None:
                summary.skipped_rows += 1
                continue
            if record.hash_key in seen:
                summary.duplicates += 1
                continue
            seen.add(record.hash_key)
            aggregated.append(record)
            summary.parsed_rows += 1
    return aggregated, summary


def _process_issue_lists(source_dir: str | None, output_dir: str, progress_area) -> list[str]:
    os.makedirs(output_dir, exist_ok=True)
    if not source_dir or not os.path.isdir(source_dir):
        progress_area.warning("未找到历史问题清单上传目录，已跳过。")
        return []
    created: list[str] = []
    files = [
        os.path.join(source_dir, name)
        for name in sorted(os.listdir(source_dir))
        if os.path.isfile(os.path.join(source_dir, name))
    ]
    if not files:
        progress_area.info("历史问题清单目录为空。")
        return created
    for file_path in files:
        name = os.path.basename(file_path)
        ext = os.path.splitext(name)[1].lower()
        progress_area.write(f"处理历史问题清单: {name}")
        try:
            if ext in SPREADSHEET_EXTENSIONS:
                records, summary = _parse_issue_list_spreadsheet(file_path)
            else:
                text_output_dir = os.path.join(output_dir, "_tmp_txt")
                os.makedirs(text_output_dir, exist_ok=True)
                txt_path = os.path.join(text_output_dir, f"{name}.txt")
                if ext in PDF_EXTENSIONS:
                    zip_bytes = _mineru_parse_pdf(file_path)
                    _zip_to_txts(zip_bytes, txt_path)
                elif ext in WORD_PPT_EXTENSIONS:
                    _unstructured_partition_to_txt(file_path, txt_path)
                else:
                    with open(file_path, "rb") as f:
                        content = f.read()
                    try:
                        text = content.decode("utf-8")
                    except UnicodeDecodeError:
                        text = content.decode("gbk", errors="ignore")
                    os.makedirs(os.path.dirname(txt_path), exist_ok=True)
                    with open(txt_path, "w", encoding="utf-8") as tmp:
                        tmp.write(text)
                if not os.path.exists(txt_path):
                    progress_area.warning(f"未生成文本，跳过: {name}")
                    continue
                with open(txt_path, "r", encoding="utf-8", errors="ignore") as f:
                    text = f.read()
                records, summary = _call_llm_for_issue_lists(text, name, progress_area)
            payload = {
                "summary": summary.to_dict(),
                "records": [record.model_dump() for record in records],
            }
            out_path = os.path.join(output_dir, f"{name}.txt")
            with open(out_path, "w", encoding="utf-8") as out_f:
                json.dump(payload, out_f, ensure_ascii=False, indent=2)
            created.append(out_path)
            progress_area.success(
                f"已生成: {os.path.basename(out_path)} — 解析{summary.parsed_rows}/{summary.total_rows}条"
            )
        except ValueError as error:
            progress_area.error(f"{name}: {error}")
        except Exception as error:
            progress_area.error(f"处理 {name} 失败: {error}")
    return created


def render_history_issues_avoidance_tab(session_id):
    if session_id is None:
        st.warning("请先登录以使用此功能。")
        return

    st.subheader("📋 历史问题规避")

    base_dirs = {
        "generated": str(CONFIG["directories"]["generated_files"]),
    }
    session_dirs = ensure_session_dirs(base_dirs, session_id)

    issue_lists_dir = session_dirs.get("history_issue_lists")
    dfmea_dir = session_dirs.get("history_dfmea")
    pfmea_dir = session_dirs.get("history_pfmea")
    cp_dir = session_dirs.get("history_cp")
    generated_root = session_dirs.get("generated_history_issues_avoidance")
    if not generated_root:
        generated_base = session_dirs.get("generated")
        if generated_base:
            generated_root = os.path.join(generated_base, "history_issues_avoidance")
            os.makedirs(generated_root, exist_ok=True)

    upload_targets = [
        {"label": "历史问题清单", "key": "issue_lists", "dir": issue_lists_dir},
        {"label": "DFMEA", "key": "dfmea", "dir": dfmea_dir},
        {"label": "PFMEA", "key": "pfmea", "dir": pfmea_dir},
        {"label": "控制计划 (CP)", "key": "cp", "dir": cp_dir},
    ]

    col_main, col_info = st.columns([2, 1])

    with col_main:
        st.markdown("请上传历史问题清单、DFMEA、PFMEA 与控制计划文件。支持 PDF、Word/PPT、Excel 等格式。")

        upload_columns = st.columns(2)
        for index, target in enumerate(upload_targets):
            column = upload_columns[index % len(upload_columns)]
            with column:
                uploaded_files = st.file_uploader(
                    f"点击上传 {target['label']}",
                    type=None,
                    accept_multiple_files=True,
                    key=f"history_upload_{target['key']}_{session_id}",
                )
                if uploaded_files:
                    handle_file_upload(uploaded_files, target["dir"])
                    st.success(f"已上传 {len(uploaded_files)} 个 {target['label']} 文件")

        st.divider()

        controls_row1 = st.columns(3)
        start_pressed = controls_row1[0].button(
            "▶️ 开始",
            key=f"history_start_{session_id}",
            use_container_width=True,
        )
        controls_row1[1].button(
            "⏸ 暂停",
            key=f"history_pause_{session_id}",
            use_container_width=True,
        )
        controls_row1[2].button(
            "🎬 演示",
            key=f"history_demo_{session_id}",
            use_container_width=True,
        )
        controls_row2 = st.columns(2)
        controls_row2[0].button(
            "⏹ 停止",
            key=f"history_stop_{session_id}",
            use_container_width=True,
        )
        controls_row2[1].button(
            "▶️ 恢复",
            key=f"history_resume_{session_id}",
            use_container_width=True,
        )

        if start_pressed:
            area = st.container()
            with area:
                if not generated_root:
                    st.error("未能初始化生成文件目录，请检查配置。")
                else:
                    progress_text = st.empty()
                    progress_bar = st.progress(0.0)
                    prompts_expander = st.expander("📝 提示与回答", expanded=False)
                    prompts_container = prompts_expander.container()
                    logs_expander = st.expander("🪵 处理日志", expanded=False)
                    logs_container = logs_expander.container()

                    def update_progress(value: float) -> None:
                        clamped = max(0.0, min(value, 1.0))
                        percent = int(round(clamped * 100))
                        progress_bar.progress(clamped)
                        progress_text.markdown(f"**当前进度：{percent}%**")

                    update_progress(0.0)
                    logs_container.info(
                        "开始处理文件：PDF 使用 MinerU，Word/PPT 使用 Unstructured，历史问题清单解析为 JSON。"
                    )
                    generated_dirs = _ensure_generated_dirs(generated_root)
                    total_created: list[str] = []
                    total_created.extend(
                        _process_issue_lists(
                            issue_lists_dir,
                            generated_dirs["issue_lists"],
                            logs_container,
                        )
                    )
                    total_created.extend(
                        _process_category(
                            "DFMEA",
                            dfmea_dir,
                            generated_dirs["dfmea"],
                            logs_container,
                        )
                    )
                    total_created.extend(
                        _process_category(
                            "PFMEA",
                            pfmea_dir,
                            generated_dirs["pfmea"],
                            logs_container,
                        )
                    )
                    total_created.extend(
                        _process_category(
                            "控制计划 (CP)",
                            cp_dir,
                            generated_dirs["cp"],
                            logs_container,
                        )
                    )
                    if total_created:
                        logs_container.success("文本解析完成。")
                        update_progress(0.1)
                        kb_id = _sync_history_kb(
                            session_id,
                            {
                                "DFMEA": generated_dirs.get("dfmea", ""),
                                "PFMEA": generated_dirs.get("pfmea", ""),
                                "控制计划": generated_dirs.get("cp", ""),
                            },
                            logs_container,
                        )
                        _warmup_history_flow(kb_id, logs_container)
                        _evaluate_history_problems(
                            generated_dirs.get("issue_lists", ""),
                            generated_dirs.get("initial_results", ""),
                            generated_dirs.get("final_results", ""),
                            kb_id,
                            session_id,
                            logs_container,
                            transcript_area=prompts_container,
                            progress_callback=update_progress,
                            progress_offset=0.1,
                            progress_span=0.85,
                        )
                        update_progress(1.0)
                    else:
                        logs_container.info("未生成任何文本文件，请确认上传内容后重试。")

    with col_info:
        st.subheader("📁 上传文件")

        clear_columns = st.columns(2)
        for index, target in enumerate(upload_targets):
            column = clear_columns[index % len(clear_columns)]
            with column:
                if st.button(
                    f"🗑️ 清空{target['label']}",
                    key=f"history_clear_{target['key']}_{session_id}",
                ):
                    try:
                        if target["dir"] and os.path.isdir(target["dir"]):
                            for name in os.listdir(target["dir"]):
                                path = os.path.join(target["dir"], name)
                                if os.path.isfile(path):
                                    os.remove(path)
                        st.success(f"已清空 {target['label']} 文件")
                        st.rerun()
                    except Exception as error:
                        st.error(f"清空失败: {error}")

        tabs = st.tabs([target["label"] for target in upload_targets])
        for tab, target in zip(tabs, upload_targets):
            with tab:
                files = _collect_files(target["dir"])
                if not files:
                    st.write("（未上传）")
                    continue
                for info in files:
                    display = _truncate_filename(info["name"])
                    with st.expander(f"📄 {display}", expanded=False):
                        col_meta, col_actions = st.columns([3, 1])
                        with col_meta:
                            st.write(f"**文件名:** {info['name']}")
                            st.write(f"**大小:** {_format_file_size(info['size'])}")
                            st.write(f"**修改时间:** {_format_timestamp(info['modified'])}")
                        with col_actions:
                            delete_key = (
                                f"history_delete_{target['key']}_{info['name'].replace(' ', '_').replace('.', '_')}_{session_id}"
                            )
                            if st.button("🗑️ 删除", key=delete_key):
                                try:
                                    os.remove(info["path"])
                                    st.success(f"已删除: {info['name']}")
                                    st.rerun()
                                except Exception as error:
                                    st.error(f"删除失败: {error}")
