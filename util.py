import os
import json
import time
import hashlib
from pathlib import Path
from typing import Dict, List

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from config import CONFIG

# --- LLM Host Resolution ---
def resolve_ollama_host(llm_backend: str) -> str:
    """Return the Ollama host URL based on selected backend key.

    Recognized keys:
    - 'ollama_127' -> CONFIG['llm']['ollama_host_127'] if present, else fallback to CONFIG['llm']['ollama_host']
    - 'ollama_9'   -> CONFIG['llm']['ollama_host_9'] if present, else fallback to CONFIG['llm']['ollama_host']
    - any other    -> CONFIG['llm']['ollama_host']
    """
    try:
        if llm_backend == "ollama_127":
            return CONFIG["llm"].get("ollama_host_127") or CONFIG["llm"]["ollama_host"]
        if llm_backend == "ollama_9":
            return CONFIG["llm"].get("ollama_host_9") or CONFIG["llm"]["ollama_host"]
    except Exception:
        pass
    return CONFIG["llm"]["ollama_host"]

# --- Login Management ---
def get_username_file():
    """Get the path to the username storage file."""
    return os.path.join(os.path.expanduser("~"), ".streamlit_username")

def save_username(username):
    """Save username to file for persistence."""
    try:
        with open(get_username_file(), 'w') as f:
            f.write(username)
        return True
    except Exception:
        return False

def load_username():
    """Load username from file if it exists."""
    try:
        if os.path.exists(get_username_file()):
            with open(get_username_file(), 'r') as f:
                return f.read().strip()
    except Exception:
        pass
    return None

# Multi-User Support (JSON-backed per-user sessions)
def get_user_sessions_dir():
    """Get the directory for storing user session files."""
    sessions_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "user_sessions")
    os.makedirs(sessions_dir, exist_ok=True)
    return sessions_dir

def get_user_session_file(username):
    """Get the path to a specific user's session file."""
    sessions_dir = get_user_sessions_dir()
    return os.path.join(sessions_dir, f"{username}_session.json")

def load_user_session(username):
    """Load user session data from JSON file if it exists."""
    try:
        session_file = get_user_session_file(username)
        if os.path.exists(session_file):
            with open(session_file, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return None

def save_user_session(username, session_data):
    """Save user session data to a JSON file."""
    try:
        session_file = get_user_session_file(username)
        with open(session_file, 'w', encoding='utf-8') as f:
            json.dump(session_data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        st.error(f"保存用户会话失败: {e}")
        return False

def create_user_session(username):
    """Create a new user session with simple username-based approach."""
    session_data = {
        "username": username,
        "login_time": time.strftime("%Y-%m-%d %H:%M:%S"),
        "last_activity": time.strftime("%Y-%m-%d %H:%M:%S"),
        "active": True
    }
    if save_user_session(username, session_data):
        return session_data
    return None

def update_user_activity(username):
    """Update the last activity timestamp for a user session."""
    session_data = load_user_session(username)
    if session_data:
        session_data["last_activity"] = time.strftime("%Y-%m-%d %H:%M:%S")
        save_user_session(username, session_data)

def deactivate_user_session(username):
    """Mark a user session as inactive (for logout)."""
    session_data = load_user_session(username)
    if session_data:
        session_data["active"] = False
        session_data["logout_time"] = time.strftime("%Y-%m-%d %H:%M:%S")
        save_user_session(username, session_data)

def get_active_users():
    """Get list of currently active users."""
    sessions_dir = get_user_sessions_dir()
    active_users = []
    if not os.path.exists(sessions_dir):
        return active_users
    for filename in os.listdir(sessions_dir):
        if filename.endswith('_session.json'):
            username = filename.replace('_session.json', '')
            session_data = load_user_session(username)
            if session_data and session_data.get('active', False):
                active_users.append(username)
    return active_users

def render_login_widget():
    """Render login widget using URL params for per-browser username and auth flag."""
    current_user = st.query_params.get("user", None)
    auth = st.query_params.get("auth", None)

    if current_user and auth == "1":
        session_data = load_user_session(current_user)
        if not session_data or not session_data.get('active', False):
            create_user_session(current_user)
        return current_user

    st.markdown("### 用户登录")
    st.markdown("使用OA用户名和密码，不用加@calb-tech.com后缀")

    prefill_username = current_user or ""

    with st.form("login_form"):
        username = st.text_input("用户名", value=prefill_username, placeholder="请输入您的用户名")
        password = st.text_input("密码", type="password", placeholder="请输入密码")

        col1, col2 = st.columns([1, 1])
        with col1:
            login_button = st.form_submit_button("登录")
        with col2:
            if current_user:
                clear_button = st.form_submit_button("清除本机用户名")
            else:
                clear_button = st.form_submit_button("清除", disabled=True)

        if login_button:
            if username.strip():
                st.session_state['logged_in'] = True
                st.session_state['username'] = username.strip()
                create_user_session(username.strip())
                st.query_params["user"] = username.strip()
                st.query_params["auth"] = "1"
                st.success(f"欢迎，{username.strip()}！")
                st.rerun()
            else:
                st.error("请输入用户名")

        if clear_button:
            try:
                if "user" in st.query_params:
                    del st.query_params["user"]
                if "auth" in st.query_params:
                    del st.query_params["auth"]
                st.success("已清除本机用户名")
                st.rerun()
            except Exception:
                st.error("清除失败")

    return None


def get_user_session_id(username):
    """Generate session ID based on username for persistence."""
    # Always derive from the provided username to avoid cross-user leakage
    return username

def get_user_session_id(username):
    """Generate session ID based on username for persistence."""
    # Always derive from the provided username to avoid cross-user leakage
    return username

# New helper to determine if a user is an admin
def is_admin(username: str) -> bool:
    admins_env = os.getenv("ADMIN_USERS", "")
    admins = {u.strip() for u in admins_env.split(",") if u.strip()}
    if not admins:
        admins = {"admin"}
    return username in admins

# --- Session Directory Utilities ---
def ensure_session_dirs(base_dirs, session_id):
    """Ensure session directories exist with lightweight session-scoped caching."""

    cache_store: Dict[str, Dict[str, str]] = st.session_state.setdefault("_session_dir_cache", {})

    resolved_roots: Dict[str, str] = {}
    for dir_type, base_dir in base_dirs.items():
        template = base_dir(session_id) if callable(base_dir) else base_dir
        if isinstance(template, (list, tuple)):
            template = os.path.join(*map(str, template))
        base_dir_str = str(template)
        if "{session_id}" in base_dir_str:
            session_dir = base_dir_str.format(session_id=session_id)
        else:
            session_dir = os.path.join(base_dir_str, str(session_id))
        resolved_roots[dir_type] = session_dir

    signature_src = "|".join(f"{key}:{resolved_roots[key]}" for key in sorted(resolved_roots))
    signature_hash = hashlib.sha1(signature_src.encode("utf-8")).hexdigest()
    cache_key = f"session_dirs::{session_id}::{signature_hash}"

    cached_dirs = cache_store.get(cache_key)
    if cached_dirs and all(os.path.isdir(path) for path in cached_dirs.values()):
        return dict(cached_dirs)

    uploads_root = str(CONFIG["directories"].get("uploads", Path(__file__).resolve().parent / "uploads"))
    os.makedirs(uploads_root, exist_ok=True)
    session_root = os.path.join(uploads_root, str(session_id))
    os.makedirs(session_root, exist_ok=True)

    session_dirs: Dict[str, str] = {}
    for dir_type, session_dir in resolved_roots.items():
        os.makedirs(session_dir, exist_ok=True)
        session_dirs[dir_type] = session_dir

    for gen_key in ("generated", "generated_files"):
        if gen_key in session_dirs:
            generated_root = session_dirs[gen_key]
            subfolders = [
                "parameters_check",
                "file_elements_check",
                "file_completeness_check",
                "history_issues_avoidance",
                "ai_agent",
            ]
            for name in subfolders:
                os.makedirs(os.path.join(generated_root, name), exist_ok=True)
            session_dirs["generated_parameters_check"] = os.path.join(generated_root, "parameters_check")
            session_dirs["generated_file_elements_check"] = os.path.join(generated_root, "file_elements_check")
            session_dirs["generated_file_completeness_check"] = os.path.join(generated_root, "file_completeness_check")
            session_dirs["generated_history_issues_avoidance"] = os.path.join(generated_root, "history_issues_avoidance")
            ai_root = os.path.join(generated_root, "ai_agent")
            for name in ("examined_txt", "initial_results", "final_results", "checkpoint", "logs"):
                os.makedirs(os.path.join(ai_root, name), exist_ok=True)
            session_dirs["generated_ai_agent"] = ai_root
            session_dirs["generated_ai_agent_examined_txt"] = os.path.join(ai_root, "examined_txt")
            session_dirs["generated_ai_agent_initial_results"] = os.path.join(ai_root, "initial_results")
            session_dirs["generated_ai_agent_final_results"] = os.path.join(ai_root, "final_results")
            session_dirs["generated_ai_agent_checkpoint"] = os.path.join(ai_root, "checkpoint")
            session_dirs["generated_ai_agent_logs"] = os.path.join(ai_root, "logs")

    def _ensure_upload_subdir(key: str, *parts: str) -> None:
        path = os.path.join(session_root, *parts)
        os.makedirs(path, exist_ok=True)
        session_dirs.setdefault(key, path)

    try:
        _ensure_upload_subdir("enterprise_standards", "enterprise_standard", "standards")
        _ensure_upload_subdir("enterprise_examined", "enterprise_standard", "examined")
        _ensure_upload_subdir("special_reference", "special_symbols", "reference")
        _ensure_upload_subdir("special_examined", "special_symbols", "examined")
        _ensure_upload_subdir("history_issue_lists", "history_issues", "issue_lists")
        _ensure_upload_subdir("history_dfmea", "history_issues", "dfmea")
        _ensure_upload_subdir("history_pfmea", "history_issues", "pfmea")
        _ensure_upload_subdir("history_cp", "history_issues", "cp")
        _ensure_upload_subdir("ai_agent_inputs", "ai_agent", "inputs")
    except Exception:
        pass

    cache_store[cache_key] = dict(session_dirs)
    st.session_state["_session_dir_cache"] = cache_store
    return session_dirs


def get_directory_refresh_token(path: str) -> float:
    """Return an mtime-based token for cache busting directory listings."""

    try:
        return os.stat(path).st_mtime
    except FileNotFoundError:
        return 0.0


@st.cache_data(show_spinner=False)
def list_directory_contents(path: str, refresh_token: float) -> tuple[Dict[str, object], ...]:
    """Return cached file metadata for ``path`` keyed by ``refresh_token``."""

    if not path or not os.path.isdir(path):
        return []

    entries: List[Dict[str, object]] = []
    for name in os.listdir(path):
        file_path = os.path.join(path, name)
        if not os.path.isfile(file_path):
            continue
        stat_result = os.stat(file_path)
        entries.append(
            {
                "name": name,
                "path": file_path,
                "size": stat_result.st_size,
                "modified": stat_result.st_mtime,
            }
        )

    return tuple(entries)


def get_file_list(folder: str) -> List[Dict[str, object]]:
    """Return a sorted list of file metadata for ``folder`` leveraging cache helpers."""

    token = get_directory_refresh_token(folder)
    entries = [dict(entry) for entry in list_directory_contents(folder, token)]
    for entry in entries:
        entry.setdefault("path", os.path.join(folder, entry["name"]))
    return sorted(entries, key=lambda item: (item["name"].lower(), item["modified"]))

# --- Structured Session State Management ---
def get_user_session(session_id, tab_name=None):
    """Get or create a structured user session with optional tab-specific state."""
    if 'user_sessions' not in st.session_state:
        st.session_state.user_sessions = {}

    if session_id not in st.session_state.user_sessions:
        st.session_state.user_sessions[session_id] = {
            'tabs': {
                'completeness': {
                    'process_started': False,
                    'analysis_completed': False,
                    'ollama_history': [],
                    'openai_history': [],
                },
                'parameters': {
                    'process_started': False,
                    'analysis_completed': False,
                    'ollama_history': [],
                    'openai_history': [],
                },
                'ai_agent': {
                    'ollama_history': [],
                    'openai_history': [],
                },
            },
            'llm_backend': 'ollama_9',
            'ollama_model': CONFIG["llm"]["ollama_model"],
            'ollama_temperature': 0.7,
            'ollama_top_p': 0.9,
            'ollama_top_k': 40,
            'ollama_repeat_penalty': 1.1,
            'ollama_num_ctx': 40001,
            'ollama_num_thread': 4,
            'openai_model': CONFIG["llm"]["openai_model"],
            'openai_temperature': 0.7,
            'openai_top_p': 1.0,
            'openai_max_tokens': 2048,
            'openai_presence_penalty': 0.0,
            'openai_frequency_penalty': 0.0,
            'openai_logit_bias': '{}',
        }

    session = st.session_state.user_sessions[session_id]

    if tab_name and tab_name in session['tabs']:
        return session['tabs'][tab_name]

    return session


def reset_user_session(session_id, tab_name=None):
    """Reset a user's session to initial state, optionally for a specific tab."""
    session = get_user_session(session_id)

    if tab_name and tab_name in session['tabs']:
        tab_session = session['tabs'][tab_name]
        tab_session['analysis_completed'] = False
        tab_session['process_started'] = False
        tab_session['ollama_history'] = []
        tab_session['openai_history'] = []
    else:
        for tab_session in session['tabs'].values():
            tab_session['analysis_completed'] = False
            tab_session['process_started'] = False
            tab_session['ollama_history'] = []
            tab_session['openai_history'] = []


def start_analysis(session_id, tab_name):
    """Start analysis for a specific tab in a user session."""
    session = get_user_session(session_id)
    if tab_name in session['tabs']:
        tab_session = session['tabs'][tab_name]
        tab_session['analysis_completed'] = False
        tab_session['process_started'] = True
        tab_session['ollama_history'] = []
        tab_session['openai_history'] = []


def complete_analysis(session_id, tab_name):
    """Mark analysis as completed for a specific tab in a user session."""
    session = get_user_session(session_id)
    if tab_name in session['tabs']:
        session['tabs'][tab_name]['analysis_completed'] = True


# --- File Upload Utility ---
def handle_file_upload(files, save_dir):
    if files:
        for file in files:
            save_path = os.path.join(save_dir, file.name)
            with open(save_path, "wb") as f:
                f.write(file.getbuffer())
        return len(files)
    return 0

# --- Prompt Generator and File Handling ---
def save_uploaded_file(uploaded_file, session_id, file_type):
    """Handle a single file upload and save to the session directory.

    Note: Renamed from handle_file_upload to avoid clashing with the
    bulk-upload helper used across tabs (files, save_dir).
    """
    if uploaded_file is not None:
        # Use the existing ensure_session_dirs function
        base_dirs = {
            'generated_files': os.path.join(os.path.dirname(os.path.abspath(__file__)), "generated_files")
        }
        session_dirs = ensure_session_dirs(base_dirs, session_id)
        session_dir = session_dirs['generated_files']
        
        file_path = os.path.join(session_dir, f"{file_type}_{uploaded_file.name}")
        
        with open(file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        
        return file_path
    return None


# --- Utility helpers (migrated from temporary scripts) ---
def json_canonical_sha256(obj) -> str:
    """Compute a canonical SHA-256 for a JSON-serializable object (sorted keys, compact)."""
    canon = json.dumps(obj, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(canon.encode("utf-8")).hexdigest().upper()

def convert_extracted_json_to_dataframe(records):
    """Convert extracted JSON format [{File, Sheet, data: [[...], ...]}, ...] to a flat DataFrame.

    Columns: File, Sheet, Row, C1..Ck
    """
    rows = []
    max_cols = 0
    for rec in records:
        file_name = rec.get("File")
        sheet_name = rec.get("Sheet")
        data = rec.get("data", []) or []
        for idx, row in enumerate(data, start=1):
            max_cols = max(max_cols, len(row))
            rows.append({
                "File": file_name,
                "Sheet": sheet_name,
                "Row": idx,
                "_cells": row,
            })
    columns = [f"C{i}" for i in range(1, max_cols + 1)]
    table_rows = []
    for r in rows:
        base = {"File": r["File"], "Sheet": r["Sheet"], "Row": r["Row"]}
        cells = r["_cells"]
        for i, col_name in enumerate(columns, start=1):
            base[col_name] = cells[i - 1] if i - 1 < len(cells) else ""
        table_rows.append(base)
    return pd.DataFrame(table_rows, columns=["File", "Sheet", "Row"] + columns)


# --- Parameter Extraction (migrated from extract_parameters.py) ---
def _column_letter_to_number(column_letter):
    if not column_letter or (isinstance(column_letter, float) and pd.isna(column_letter)):
        return None
    result = 0
    for char in str(column_letter).strip():
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result

def _parse_column_list(column_string):
    if not column_string or (isinstance(column_string, float) and pd.isna(column_string)):
        return []
    columns = [col.strip() for col in str(column_string).split(',')]
    return [_column_letter_to_number(col) for col in columns if col]

def _parse_row_list(row_string):
    if not row_string or (isinstance(row_string, float) and pd.isna(row_string)):
        return []
    rows = [row.strip() for row in str(row_string).split(',')]
    return [int(row) for row in rows if row.isdigit()]

def _find_sheet_in_workbook(workbook, target_sheet_name):
    target_clean = str(target_sheet_name).strip()
    if target_sheet_name in workbook.sheetnames:
        return target_sheet_name
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip() == target_clean:
            return sheet_name
    return None

def _extract_sheet_data(worksheet, header_rows, process_column, parameter_columns, value_columns):
    extracted_data = []
    all_columns = []
    if process_column:
        all_columns.append(process_column)
    all_columns.extend(parameter_columns)
    all_columns.extend(value_columns)
    all_columns = sorted(list(set(all_columns)))

    remark_row = None
    for row_num in range(worksheet.max_row, max(header_rows) if header_rows else 0, -1):
        for col_num in range(1, min(worksheet.max_column + 1, 6)):
            cell_value = worksheet.cell(row=row_num, column=col_num).value
            if cell_value and str(cell_value).strip() == "备注":
                remark_row = row_num
                break
        if remark_row:
            break

    for row_num in header_rows:
        if row_num <= worksheet.max_row:
            row_data = []
            for col_num in all_columns:
                if col_num <= worksheet.max_column:
                    cell_value = worksheet.cell(row=row_num, column=col_num).value
                    row_data.append("" if cell_value == "/" else (cell_value if cell_value is not None else ""))
                else:
                    row_data.append("")
            extracted_data.append(row_data)

    if header_rows:
        max_header_row = max(header_rows)
        end_row = remark_row if remark_row else worksheet.max_row
        for row_num in range(max_header_row + 1, end_row):
            row_data = []
            for col_num in all_columns:
                if col_num <= worksheet.max_column:
                    cell_value = worksheet.cell(row=row_num, column=col_num).value
                    row_data.append("" if cell_value == "/" else (cell_value if cell_value is not None else ""))
                else:
                    row_data.append("")
            extracted_data.append(row_data)

    filtered_data = []
    for row_data in extracted_data:
        if any(cell_value and str(cell_value).strip() for cell_value in row_data):
            filtered_data.append(row_data)
    return filtered_data, all_columns

def _remove_fubiao_tables(data):
    filtered_data = []
    i = 0
    while i < len(data):
        row_data = data[i]
        is_fubiao_start = any(
            cell_value and isinstance(cell_value, str) and "附表(" in str(cell_value)
            for cell_value in row_data
        )
        if is_fubiao_start:
            while i < len(data):
                row_data = data[i]
                if (
                    row_data and len(row_data) > 0 and row_data[0] and isinstance(row_data[0], str) and row_data[0].startswith("File:")
                ):
                    break
                i += 1
            continue
        filtered_data.append(row_data)
        i += 1
    return filtered_data

def extract_parameters_to_json(cp_session_dir, output_json_path, config_csv_path=None):
    cp_dir = Path(cp_session_dir)
    if not cp_dir.exists():
        raise FileNotFoundError(f"CP session dir not found: {cp_session_dir}")
    if config_csv_path is None:
        config_csv_path = str(cp_dir / "excel_sheets.csv")
    if not os.path.exists(config_csv_path):
        raise FileNotFoundError(
            f"Configuration CSV not found: {config_csv_path}. Place 'excel_sheets.csv' under {cp_session_dir}."
        )
    df = pd.read_csv(config_csv_path)
    false_sheets = df[df['skip'].astype(str).str.upper() == 'FALSE']
    json_data = []
    total_rows = 0
    for _, config in false_sheets.iterrows():
        file_name = str(config['file'])
        sheet_name = str(config['sheet']).strip()
        file_path = cp_dir / file_name
        if not file_path.exists():
            continue
        try:
            workbook = openpyxl.load_workbook(file_path)
            actual_sheet_name = _find_sheet_in_workbook(workbook, sheet_name)
            if not actual_sheet_name:
                continue
            worksheet = workbook[actual_sheet_name]
            header_rows = _parse_row_list(config.get('header_rows'))
            process_column = _column_letter_to_number(config.get('process_number_column')) if config.get('process_number_column') else None
            parameter_columns = _parse_column_list(config.get('parameter_columns'))
            value_columns = _parse_column_list(config.get('value_columns'))
            data, _ = _extract_sheet_data(worksheet, header_rows, process_column, parameter_columns, value_columns)
            if not data:
                continue
            filtered_data = _remove_fubiao_tables(data)
            total_rows += len(filtered_data)
            sheet_data = {"File": file_name, "Sheet": sheet_name, "data": filtered_data}
            json_data.append(sheet_data)
        except Exception:
            continue
    output_path = Path(output_json_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as json_file:
        json.dump(json_data, json_file, ensure_ascii=False, indent=2)
    return {"output": str(output_path), "sheets": len(json_data), "rows": total_rows, "config": config_csv_path}
