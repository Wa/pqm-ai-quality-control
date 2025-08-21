"""
Parameter extraction utilities for the Parameters Check tab.

Adapted from the standalone script provided by the user (extract_data_as_json.py).
This module exposes a single entry point:

    extract_parameters_to_json(cp_session_dir, output_json_path, config_csv_path=None)

Behavior:
- Reads a CSV configuration (excel_sheets.csv) describing which sheets/columns to parse
- Loads workbooks under cp_session_dir
- Extracts rows/columns per config and writes consolidated JSON to output_json_path
"""

from __future__ import annotations

import os
import json
from pathlib import Path
from typing import List, Optional, Tuple

import pandas as pd
import openpyxl


def column_letter_to_number(column_letter: Optional[str]) -> Optional[int]:
    if not column_letter or pd.isna(column_letter):
        return None
    result = 0
    for char in str(column_letter).strip():
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result


def parse_column_list(column_string: Optional[str]) -> List[int]:
    if not column_string or pd.isna(column_string):
        return []
    columns = [col.strip() for col in str(column_string).split(',')]
    return [column_letter_to_number(col) for col in columns if col]


def parse_row_list(row_string: Optional[str]) -> List[int]:
    if not row_string or pd.isna(row_string):
        return []
    rows = [row.strip() for row in str(row_string).split(',')]
    return [int(row) for row in rows if row.isdigit()]


def find_sheet_in_workbook(workbook: openpyxl.Workbook, target_sheet_name: str) -> Optional[str]:
    target_clean = target_sheet_name.strip()
    if target_sheet_name in workbook.sheetnames:
        return target_sheet_name
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip() == target_clean:
            return sheet_name
    return None


def extract_sheet_data(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    header_rows: List[int],
    process_column: Optional[int],
    parameter_columns: List[int],
    value_columns: List[int],
) -> Tuple[List[List[str]], List[int]]:
    extracted_data: List[List[str]] = []
    all_columns: List[int] = []

    if process_column:
        all_columns.append(process_column)
    all_columns.extend(parameter_columns)
    all_columns.extend(value_columns)
    all_columns = sorted(list(set(all_columns)))

    remark_row = None
    for row_num in range(worksheet.max_row, max(header_rows) if header_rows else 0, -1):
        for col_num in range(1, min(worksheet.max_column + 1, 6)):
            cell_value = worksheet.cell(row=row_num, column=col_num).value
            if cell_value and str(cell_value).strip() == "备注":
                remark_row = row_num
                break
        if remark_row:
            break

    for row_num in header_rows:
        if row_num <= worksheet.max_row:
            row_data: List[str] = []
            for col_num in all_columns:
                if col_num <= worksheet.max_column:
                    cell_value = worksheet.cell(row=row_num, column=col_num).value
                    row_data.append("" if cell_value == "/" else (cell_value if cell_value is not None else ""))
                else:
                    row_data.append("")
            extracted_data.append(row_data)

    if header_rows:
        max_header_row = max(header_rows)
        end_row = remark_row if remark_row else worksheet.max_row
        for row_num in range(max_header_row + 1, end_row):
            row_data: List[str] = []
            for col_num in all_columns:
                if col_num <= worksheet.max_column:
                    cell_value = worksheet.cell(row=row_num, column=col_num).value
                    row_data.append("" if cell_value == "/" else (cell_value if cell_value is not None else ""))
                else:
                    row_data.append("")
            extracted_data.append(row_data)

    filtered_data: List[List[str]] = []
    for row_data in extracted_data:
        if any(cell_value and str(cell_value).strip() for cell_value in row_data):
            filtered_data.append(row_data)
    return filtered_data, all_columns


def remove_fubiao_tables(data: List[List[str]]) -> List[List[str]]:
    filtered_data: List[List[str]] = []
    i = 0
    while i < len(data):
        row_data = data[i]
        is_fubiao_start = any(
            cell_value and isinstance(cell_value, str) and "附表(" in str(cell_value)
            for cell_value in row_data
        )
        if is_fubiao_start:
            while i < len(data):
                row_data = data[i]
                if (
                    row_data
                    and len(row_data) > 0
                    and row_data[0]
                    and isinstance(row_data[0], str)
                    and row_data[0].startswith("File:")
                ):
                    break
                i += 1
            continue
        filtered_data.append(row_data)
        i += 1
    return filtered_data


def extract_parameters_to_json(
    cp_session_dir: str,
    output_json_path: str,
    config_csv_path: Optional[str] = None,
) -> dict:
    """Extract parameters per CSV config and write a consolidated JSON file.

    Returns a summary dict with counts and output path.
    """
    cp_dir = Path(cp_session_dir)
    if not cp_dir.exists():
        raise FileNotFoundError(f"CP session dir not found: {cp_session_dir}")

    if config_csv_path is None:
        config_csv_path = str(cp_dir / "excel_sheets.csv")
    if not os.path.exists(config_csv_path):
        raise FileNotFoundError(
            f"Configuration CSV not found: {config_csv_path}. Place 'excel_sheets.csv' under {cp_session_dir}."
        )

    df = pd.read_csv(config_csv_path)
    false_sheets = df[df['skip'].astype(str).str.upper() == 'FALSE']

    json_data = []
    total_rows = 0

    for _, config in false_sheets.iterrows():
        file_name = str(config['file'])
        sheet_name = str(config['sheet']).strip()

        file_path = cp_dir / file_name
        if not file_path.exists():
            # Skip silently but record warning-like entry
            continue

        try:
            workbook = openpyxl.load_workbook(file_path)
            actual_sheet_name = find_sheet_in_workbook(workbook, sheet_name)
            if not actual_sheet_name:
                continue

            worksheet = workbook[actual_sheet_name]
            header_rows = parse_row_list(config.get('header_rows'))
            process_column = (
                column_letter_to_number(config.get('process_number_column'))
                if config.get('process_number_column')
                else None
            )
            parameter_columns = parse_column_list(config.get('parameter_columns'))
            value_columns = parse_column_list(config.get('value_columns'))

            data, _ = extract_sheet_data(
                worksheet, header_rows, process_column, parameter_columns, value_columns
            )
            if not data:
                continue

            filtered_data = remove_fubiao_tables(data)
            total_rows += len(filtered_data)

            sheet_data = {
                "File": file_name,
                "Sheet": sheet_name,
                "data": filtered_data,
            }
            json_data.append(sheet_data)

        except Exception:
            # Skip on processing error for this file/sheet
            continue

    output_path = Path(output_json_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as json_file:
        json.dump(json_data, json_file, ensure_ascii=False, indent=2)

    return {
        "output": str(output_path),
        "sheets": len(json_data),
        "rows": total_rows,
        "config": config_csv_path,
    }



